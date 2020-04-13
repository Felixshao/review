import unittest
import requests, hashlib
import ssl
import warnings
import json
import xlrd
import openpyxl
import paramunittest
from xlutils3.copy import copy
from openpyxl.styles import Font, colors


class inter_case(unittest.TestCase):

    @classmethod
    def setUpClass(cls):

        warnings.filterwarnings('ignore')   # 忽略warning警告
        #
        # cls.log = open('.\\report\\inter_log.txt', mode='w')
        # cls.log.truncate()
        # cls.log.close()
        #
        # cls.log = open('.\\report\\inter_log.txt', mode='a+', encoding='utf-8')


    @classmethod
    def tearDownClass(cls):
        # cls.log.close()
        pass

    """boos搜索接口"""
    def test1_boosSearch(self):

        url = 'https://www.zhipin.com/wapi/zpgeek/autocomplete/query.json'
        params = {'query': u'测试'}

        # requests.get(),get方式请求接口，params为参数，verify=False移除ssl证书认证，不移除访问https类的接口会报错
        search = requests.get(url=url, params=params, verify=False)
        print(search.url)
        dict = json.loads(search.text)
        # print(dict['code'], dict)
        print(dict)

        # assertEqual()断言方法，传入值（真实的值，期待的值，msg=''） 两值相等为true，否则为false，打印出msg
        # self.assertEqual(dict['code'], 0)
        # print(dict['code'], dict)

    def test11_boosSearch2(self):

        url = 'https://www.zhipin.com/c101280600/'
        params = {'query': '自动化测试', 'period': 1, 'ka': 'sel-scale-1'}
        # requests.get(),get方式请求接口，params为参数，verify=False移除ssl证书认证，不移除访问https类的接口会报错
        # allow_redirects=False参数，阻止接口跟随重定向
        search = requests.get(url=url, params=params, verify=False, allow_redirects=False)
        print(search.url)
        print(search.status_code)
        new_url = 'https://www.zhipin.com' + search.headers['Location']  # search.headers['Location']获取重定的url
        print(new_url)
        search = requests.get(url=new_url, verify=False, allow_redirects=False)
        print(search.content.decode('utf-8'))
        a = '%3D%25E8%2587%25AA%25E5%258A%25A8%25E5%258C%2596%25E6%25B5%258B%25E8%25AF%2595%26period%3D2%26ka%3D'

    def test2_boosLogin(self):

        login_url = 'https://login.zhipin.com/wapi/zppassport/qrcode/dispatcher'
        params = {'qrId': 'bosszp-b4e0fae4-0482-420e-973e-004458e06a95'}

        login = requests.get(login_url, params=params, verify=False)
        login_dict = json.loads(login.text)
        cookies = login.cookies
        print(login_dict, '\n', cookies)

    def test3_readExcel(self):
        """
        引入xlrd包读取case，
        引入requests包，请求接口
        :return:
        """
        file_name = '..\\file\\inter.xlsx'

        # xlrd.open_workbook(文件路径)方法读取文件内容；sheet_by_index()选择文件第几张表，参数为数字
        table = xlrd.open_workbook(file_name)
        data = table.sheet_by_index(0)
        row = data.nrows        # 获取行数
        col = data.ncols        # 获取列数
        all_list = []
        for i in range(1, row):
            list = []
            for j in range(1, col):
                if data.cell_value(i, j) != '':
                    # 按行逐个读取数据传入列表中
                    list.append(data.cell_value(i, j).replace('\n', ''))
            # 每读取一行存入一个新的列表中，存入所有接口数据
            all_list.append(list)

        print(all_list)
        boss_search = all_list[0]        # 选择需要操作的接口
        boss_search[2] = eval(boss_search[2])   # 读取的参数为str类型，使用eval()方法转为dict类型
        # 判断请求是get还是post方法
        if boss_search[1] == 'get':
            search = requests.get(boss_search[0], params=boss_search[2], verify=False)
        else:
            search = requests.post(boss_search[0], params=boss_search[2], verify=False)
        print(search.url)

        # 写入结果
        table2 = copy(table)
        sheet = table2.get_sheet(0)

        # 判断响应信息是json报文还是网页报文
        search.dict = json.loads(search.text)
        try:
            self.assertEqual(search.dict['code'], 0)
            sheet.write(1, 4, '通过')
        except BaseException as c:
            sheet.write(1, 4, '不通过')
            erron = 'test3_readExcel: 参数为 "' + boss_search[2]['query'] + ' "时:' + '报错，\n错误为：' + search.text + '\n\n'
            # self.log.write(erron)

        table2.save('..\\file\\inter.xlsx')

    # def test4_writeExcel(self):
    #
    #     table3 = load_workbook('..\\file\\test.xlsx')
    #     print(type(table3), table3.sheetnames)
    #     table4 = table3.active
    #     table4.cell(3, 3, '哈哈1')
    #     table3.active = 4
    #     table3.save('..\\file\\test.xlsx')

    """1.1获取接口并写入结果"""
    def test10_csdnLogin(self):

        cookies = {
            'laravel_session': 'eyJpdiI6Iml2RWpiNGJuamI2SDNCZVRHZ1dnZEE9PSIsInZhbHVlIjoiTU0zaXhlT3RmVlJibG1VUEpuVzJMc'
                               '2d5aDFiZFBqSTdPRFJxenU3cnVkN01DZnZQRVVybVRUSFBac2Y5dWN2TVwvdjZBaFRZVEhOaUxTY25yOHFyR0'
                               'pBPT0iLCJtYWMiOiIyY2NmMjBhYzY5NTJjNjE3MmQ5MDYxODdhYWFiMTc1MDkwNTRkMWYzYzFlYjk4YmFlZTN'
                               'mZGMxMzVjMTEzYTBjIn0%3D',
            'b_user_token': 'token_5dbfe87e01a375MQBHPraMkqNJXwclQfL'
        }

        file = '..\\file\\inite2.xlsx'        # 接口文件地址
        table = xlrd.open_workbook(file).sheet_by_index(1)      # 打开文件获取第一个表数据
        table_title = table.name            # 获取当前sheet名字
        row = table.nrows       # 行数
        col = table.ncols       # 列数
        all_data = []
        # 读取接口信息并存放在list中
        for i in range(1, row):
            data = []
            for j in range(1, col):
                data.append(table.cell_value(i, j).replace('\n', ''))
            all_data.append(data)

        # 请求接口并断言结果
        result_list = []  # 存放结果的list

        for data in all_data:
            data[1] = eval(data[1])     # 将参数转化为字典格式
            if data[2] == 'get':
                reques = requests.get(data[0], params=data[1], cookies=cookies, verify=False)
                result = json.loads(reques.text)    # 获取响应结果并转化为字
                try:
                    self.assertEqual(result['code'], 0, msg='erron')
                    result_list.append('pass')
                except BaseException as b:
                    print(b)
                    result_list.append('erron')

            else:
                reques = requests.post(data[0], params=data[1], cookies=cookies, verify=False)
                result = json.loads(reques.text)     # 获取响应结果并转化为字典
                try:
                    self.assertEqual(int(result['code']), 0, msg='erron')
                    result_list.append('pass')
                except BaseException as b:
                    print(b)
                    result_list.append('erron')

        # 结果写入excel表格中
        result_list = ['pass', 'pass']
        excel = openpyxl.load_workbook(file)    # 生成一个已存在的Workbook对象
        sheet = excel.get_sheet_by_name(table_title)   # 获取需要操作的表格
        row = 2
        for result in result_list:

            sheet.cell(row, 5, value=result)      # 在指定格子写入数据
            row += 1
        excel.save(file)    # 保存文件，运行时需关闭文件，打开会报错
        print('写入成功，请查看！')

    """获取接口数据请求并断言"""
    def test11_getExcel(self):
        """
        打开excel取得接口信息
        all_inter['name']:接口名称
         all_inter['value']:接口具体信息
        """

        file = 'D:\\study\\python\\review\\file\\inite2.xlsx'

        cookies = {'laravel_session': 'eyJpdiI6IkJhSjlSbVdsWVVFM3g0Y2JTNDYydHc9PSIsInZhbHVlIjoiQzZXVGtScU1zZCtvSElva3h'
                                      'Rdk13N1dcLzI1Z2lZYytsWkNVaEUwQlZSM2FzTm0yNkFpYWk2RW52Ykk2V2tpcjFxU3NWUWxNdUNcL2'
                                      'pZNU45XC9hSjFUQ0E9PSIsIm1hYyI6ImI4ZTkxYTA5YzUwNDA0OGFlNzMwZDI1OTgyNDY0MDE1ZDhiO'
                                      'GVjYWU1NjYyZTEwMDM1ZjAyMjE0MjkyZGYyMDcifQ%3D%3D',
                   'b_user_token': 'token_5dc415c93a20e3DWTlI1bXQFZnJ2Pz7pS'}

        table = xlrd.open_workbook(file).sheet_by_index(0)  # 打开文件的第一个表格
        sheet_name = table.name     # 获取第一个sheet的name
        row = table.nrows           # 获取行
        col = table.ncols           # 获取列
        all_inter = {}
        for i in range(1, row):
            inter = []             # 临时列，用来存放一组接口数据
            for j in range(0, col-1):
                inter.append(table.cell_value(i, j).replace('\n', ''))  # 逐列获取数据插入list中，有换行符替换成空
            all_inter[inter[0]] = inter[1:]     # 一个接口数据存入字典中

        print(all_inter)
        event = []
        for i in all_inter:
            print('{0}接口开始测试！'.format(i))
            url = all_inter[i][0]
            params = eval(all_inter[i][1])
            reques = all_inter[i][2]

            if reques == 'get':
                try:
                    inter = requests.get(url, params=params, cookies=cookies, verify=False)
                    result = json.loads(inter.text)
                    print(result)
                    self.assertEqual(int(result['code']), 0, msg='erron，响应结果：' + str(result))
                    flag = 'pass'
                except:
                    flag = 'erron'

            elif reques == 'post':
                try:
                    inter = requests.post(url, data=params, cookies=cookies, verify=False)
                    result = json.loads(inter.text)
                    self.assertEqual(int(result['code']), 0, msg='erron，响应结果：' + str(result))
                    flag = 'pass'
                except:
                    flag = 'erron'
            else:
                flag = 'erron, 无此请求方式！'

            event.append(flag)  # 记录接口运行结果
        print(event)

        table2 = openpyxl.load_workbook(file)   # 生成一个已存在的workbook对象
        sheet = table2.get_sheet_by_name('xiaoe2')    # 根据表名选择活动表
        j = 2
        for i in event:
            sheet.cell(j, 5, i)
            j += 1
        table2.save(file)

    def test12_easy_getExcet(self):
        file = 'D:\\study\\python\\review\\file\\inite2.xlsx'

        table = xlrd.open_workbook(file)
        sheet = table.sheet_by_name('xiaoe1')
        all_inter = [[sheet.cell_value(i, j).replace('\n', '') for j in range(1, sheet.ncols)]
                     for i in range(1, sheet.nrows)]

        print(all_inter)

        # dict = {'name1': '123'}
        # dict2 = {'value1': '456'}
        # dict3 = {'value1': '456'}
        # dict.update(dict2)  # 更新dict（插入数据）
        # dict.pop('name1')   # 去除字典指定数据
        # print(dict)

        # list = [10, 20, 5, 14, 16, 22, 11, 55, 30, 22, 15, 19]
        # tem = 0
        # # 冒泡排序
        # for i in range(len(list)-1):
        #     ex_flag = False
        #     for j in range(len(list)-1-i):
        #         if list[j] < list[j+1]:
        #             list[j], list[j+1] = list[j+1], list[j]     # 两值相互交换
        #             print(list[j], list[j+1])
        #             ex_flag = True
        #     tem += 1
        #     if ex_flag:
        #         print(tem, list)    # 检测冒泡排序循环次数

    """读写txt文件"""
    def test13_ReadWriteText(self):

        file = 'D:\\study\\python\\review\\file\\inter.txt'
        txt = open(file)    # 打开文件，只读模式
        # data = txt.read()   # 读取txt文件数据存入字符串
        data = txt.readlines()  # 获取所有数据存入list
        all_list = [row.replace('\n', '').split(' ') for row in data]   # 切割和剔除换行符后按list存放数据
        print(all_list, '\n', type(all_list))
        txt.close()  # 关闭文件
        # 在list中每行末尾参数结果
        for i in range(1, len(all_list)):
            try:
                all_list[i][4] = 'erron'
            except:
                all_list[i].append('erron')

        # 写入方式打开文件
        txt = open(file, 'w')
        # 把list中的数据逐个覆盖写入文件中，注意格式
        for i in all_list:
            for s in i:
                txt.write(s)
                if s != i[4]:
                    txt.write(' ')
            txt.write('\n')

        # 关闭文件
        txt.close()

    """设置session会话保持"""
    def test14_keepSession(self):

        s = requests.Session()  # 定义一个全局session
        print(s, type(s))

        url = 'https://admin.xiaoe-tech.com/punch_card/get_clock_user_list'
        params = {'page': '1', 'page_size': '15', 'search_content': '你好'}
        cookies = {'b_user_token': 'token_5dc4df81c5730o3p6PIivUHAj3owBBA9m',
                   'laravel_session': 'eyJpdiI6IjlGRHhpRFpzXC9TaTRZUll3eEM4S2JBPT0iLCJ2YWx1ZSI6IjVLaXlLb1M0WDBNQ1IxVzR'
                                      'UOStibDhHb2IxSXpUbVlUTVNqUmtUYXpNZWo1ajRJK29ZUEVUSGFBS29UZnhXYnl6R283d2R3RFpcL1'
                                      'QyWkhMbTZndk8yQT09IiwibWFjIjoiNjc0OTFjNTA0NmNjY2Q2NDM2NWFjMTVlNGViNDFkMDViNDBhN'
                                      'jkwMzNkMTA3MmUwZjdmNzA1YzAwNDgyZWY2YiJ9'}
        car = s.post(url, data=params, cookies=cookies, verify=False)
        print(car.text)

        url1 = 'https://admin.xiaoe-tech.com/new/customerList'
        params1 = {'is_pay': '0', 'search': 'f', 'page': '1', 'ruler': ''}

        user = s.get(url1, params=params1, verify=False)    # 通过session对象请求
        print(user.text)

    """传入md5加密参数"""
    def test15_md5_parameter(self):

        url = 'http://127.0.0.1:8882/login'
        parame = {'name': 'xiao', 'pwd': '123456'}
        for i in parame:
            h1 = hashlib.md5()
            h1.update(parame[i].encode(encoding='utf-8'))
            parame[i] = h1.hexdigest()
        print(parame)
        login = requests.get(url, params=parame, verify=False)
        print(login.text)

    def test16_stock(self):
        url = 'http://qd.10jqka.com.cn/quote.php'
        params = {'cate': 'real', 'type': 'stock', 'return': 'json', 'callback': 'showStockData', 'code': 601698}
        # headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
        #                          'Chrome/74.0.3729.157 Safari/537.36'}
        stock_data = requests.get(url, params=params)
        print(stock_data.text)


if __name__ == '__main__':
    pass
    # base.testCase()




















